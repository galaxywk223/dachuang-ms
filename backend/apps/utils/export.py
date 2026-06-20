
import re
import zipfile
from io import BytesIO
from datetime import datetime
from pathlib import Path
import openpyxl  # type: ignore[import-untyped]
from django.utils.timezone import localtime


SAFE_ARCNAME_SEGMENT_RE = re.compile(r'[\x00-\x1f<>:"\\|?*]+')


def _safe_zip_segment(value):
    safe_parts = []
    raw = str(value or "").replace("\\", "/")
    for part in raw.split("/"):
        part = SAFE_ARCNAME_SEGMENT_RE.sub("_", part).strip()
        part = part.strip(". ")
        if part and part != "..":
            safe_parts.append(part)
    return "_".join(safe_parts)


def safe_zip_arcname(value, fallback="file"):
    """
    Return a relative, traversal-free ZIP entry name.
    """
    raw = str(value or "").replace("\\", "/")
    segments = []
    for segment in raw.split("/"):
        segment = _safe_zip_segment(segment)
        if not segment or segment == "..":
            continue
        segments.append(segment)
    if not segments:
        return fallback
    return "/".join(segments)


def safe_zip_path(*segments, fallback="file"):
    safe_segments = []
    for segment in segments:
        safe_segment = _safe_zip_segment(segment)
        if safe_segment and safe_segment != "..":
            safe_segments.append(safe_segment)
    if not safe_segments:
        return fallback
    return "/".join(safe_segments)

def generate_excel(data, headers, filename="export.xlsx"):
    """
    Generate an Excel file from a list of dictionaries.
    :param data: List of dictionaries or objects
    :param headers: Dictionary mapping field names to column headers {field: header}
    :param filename: Filename for the download
    :return: BytesIO object containing the Excel file
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export Data"

    # Write headers
    header_keys = list(headers.keys())
    for col_idx, header in enumerate(headers.values(), 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data
    for row_idx, item in enumerate(data, 2):
        for col_idx, field in enumerate(header_keys, 1):
            value = getattr(item, field, None) if hasattr(item, field) else item.get(field)
            
            # Handle Choice Fields (get display value)
            if hasattr(item, f"get_{field}_display"):
                 value = getattr(item, f"get_{field}_display")()
            
            # Handle standard fields
            if isinstance(value, datetime):
                value = localtime(value).strftime("%Y-%m-%d %H:%M:%S")
            elif value is None:
                value = ""
            
            ws.cell(row=row_idx, column=col_idx, value=str(value))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_zip(files, filename="attachments.zip"):
    """
    Generate a ZIP file from a list of file paths or file fields.
    :param files: List of tuples (file_path_or_file_field, arcname_in_zip)
    :param filename: Filename for the download
    :return: BytesIO object containing the ZIP file
    """
    output = BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_source, arcname in files:
            safe_arcname = safe_zip_arcname(arcname)
            if hasattr(file_source, "open"):
                with file_source.open("rb") as source:
                    zf.writestr(safe_arcname, source.read())
            else:
                file_path = Path(file_source)
                if file_path.exists():
                    zf.write(file_path, safe_arcname)
    
    output.seek(0)
    return output
