from decimal import Decimal
from io import BytesIO
from unittest.mock import patch
import zipfile

import openpyxl  # type: ignore[import-untyped]
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from rest_framework.test import APIClient

from apps.dictionaries.models import DictionaryItem, DictionaryType
from apps.projects.models import (
    Project,
    ProjectAchievement,
    ProjectAdvisor,
    ProjectArchive,
    ProjectMember,
)
from apps.system_settings.models import ProjectBatch
from apps.users.models import Role


User = get_user_model()


class StorageBackedFile:
    name = "remote/proposal.pdf"

    @property
    def path(self):
        raise NotImplementedError("Remote storage has no local path")

    def open(self, mode="rb"):
        return BytesIO(b"proposal-content")


class ProjectAdminImportPermissionsTestCase(TestCase):
    def setUp(self):
        self.level2_role, _ = Role.objects.get_or_create(
            code="LEVEL2_ADMIN",
            defaults={"name": "学院管理员", "scope_dimension": "COLLEGE"},
        )
        self.level1_role, _ = Role.objects.get_or_create(
            code="LEVEL1_ADMIN",
            defaults={"name": "校级管理员", "scope_dimension": "SCHOOL"},
        )
        self.level2_role.scope_dimension = "COLLEGE"
        self.level2_role.is_active = True
        self.level2_role.save(update_fields=["scope_dimension", "is_active"])
        self.level1_role.scope_dimension = "SCHOOL"
        self.level1_role.is_active = True
        self.level1_role.save(update_fields=["scope_dimension", "is_active"])
        self.custom_school_role, _ = Role.objects.get_or_create(
            code="SCHOOL_PROJECT_DIRECTOR",
            defaults={"name": "校级项目主管", "scope_dimension": "SCHOOL"},
        )
        self.custom_school_role.scope_dimension = "SCHOOL"
        self.custom_school_role.is_active = True
        self.custom_school_role.save(update_fields=["scope_dimension", "is_active"])

        self.level2_admin = User.objects.create_user(
            username="level2_import_admin",
            password="password123",
            role_fk=self.level2_role,
            real_name="学院管理员",
            employee_id="L20001",
            college="计算机学院",
        )
        self.level1_admin = User.objects.create_user(
            username="level1_project_admin",
            password="password123",
            role_fk=self.level1_role,
            real_name="校级管理员",
            employee_id="L10001",
        )
        self.custom_school_admin = User.objects.create_user(
            username="custom_school_project_admin",
            password="password123",
            role_fk=self.custom_school_role,
            real_name="校级项目主管",
            employee_id="CS1001",
        )
        self.student = User.objects.create_user(
            username="project_leader",
            password="password123",
            role_fk=Role.objects.get(code="STUDENT"),
            real_name="项目负责人",
            employee_id="S90001",
            college="计算机学院",
        )
        self.batch = ProjectBatch.objects.create(
            name="2026",
            year=2026,
            code="B2026",
            status=ProjectBatch.STATUS_ACTIVE,
            is_active=True,
            is_current=True,
        )

        self.client = APIClient()
        self.client.force_authenticate(user=self.level2_admin)

    def _achievement_type(self):
        dict_type, _ = DictionaryType.objects.get_or_create(
            code="achievement_type",
            defaults={"name": "成果类型"},
        )
        item, _ = DictionaryItem.objects.get_or_create(
            dict_type=dict_type,
            value="PAPER",
            defaults={"label": "论文", "sort_order": 1},
        )
        return item

    def test_level2_admin_cannot_use_legacy_history_import(self):
        response = self.client.post(
            "/api/v1/projects/admin/manage/import-history/",
            {},
            format="multipart",
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.data["message"], "无权限导入历史项目")

    def test_admin_project_list_tolerates_invalid_pagination(self):
        Project.objects.create(
            project_no="DC20260001",
            title="分页参数测试项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.get(
            "/api/v1/projects/admin/manage/",
            {"page": "invalid", "page_size": "1000"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["data"]["page"], 1)
        self.assertEqual(response.data["data"]["page_size"], 100)
        self.assertEqual(response.data["data"]["total"], 1)

    def test_custom_school_admin_sees_all_college_projects(self):
        Project.objects.create(
            project_no="DC20260009",
            title="本学院项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        other_student = User.objects.create_user(
            username="custom_school_other_college_leader",
            password="password123",
            role_fk=Role.objects.get(code="STUDENT"),
            real_name="其他学院负责人",
            employee_id="S90004",
            college="数学学院",
        )
        Project.objects.create(
            project_no="DC20260010",
            title="其他学院项目",
            leader=other_student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        self.client.force_authenticate(user=self.custom_school_admin)

        response = self.client.get("/api/v1/projects/admin/manage/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["data"]["total"], 2)

    def test_legacy_level1_admin_without_scope_sees_all_college_projects(self):
        self.level1_role.scope_dimension = None
        self.level1_role.save(update_fields=["scope_dimension"])
        self.level1_admin.refresh_from_db()
        Project.objects.create(
            project_no="DC20260014",
            title="旧校级本学院项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        other_student = User.objects.create_user(
            username="legacy_school_other_college_leader",
            password="password123",
            role_fk=Role.objects.get(code="STUDENT"),
            real_name="旧校级其他学院负责人",
            employee_id="S90006",
            college="数学学院",
        )
        Project.objects.create(
            project_no="DC20260015",
            title="旧校级其他学院项目",
            leader=other_student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.get("/api/v1/projects/admin/manage/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["data"]["total"], 2)

    def test_custom_school_admin_sees_all_admin_achievements(self):
        achievement_type = self._achievement_type()
        project = Project.objects.create(
            project_no="DC20260012",
            title="本学院成果项目",
            leader=self.student,
            status=Project.ProjectStatus.CLOSED,
            year=2026,
            batch=self.batch,
        )
        other_student = User.objects.create_user(
            username="custom_school_achievement_leader",
            password="password123",
            role_fk=Role.objects.get(code="STUDENT"),
            real_name="其他学院成果负责人",
            employee_id="S90005",
            college="数学学院",
        )
        other_project = Project.objects.create(
            project_no="DC20260013",
            title="其他学院成果项目",
            leader=other_student,
            status=Project.ProjectStatus.CLOSED,
            year=2026,
            batch=self.batch,
        )
        ProjectAchievement.objects.create(
            project=project,
            achievement_type=achievement_type,
            title="本学院成果",
            description="D1",
            authors="A1",
        )
        ProjectAchievement.objects.create(
            project=other_project,
            achievement_type=achievement_type,
            title="其他学院成果",
            description="D2",
            authors="A2",
        )
        self.client.force_authenticate(user=self.custom_school_admin)

        response = self.client.get("/api/v1/projects/admin/achievements/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 2)

    def test_legacy_level1_admin_without_scope_sees_all_admin_achievements(self):
        self.level1_role.scope_dimension = None
        self.level1_role.save(update_fields=["scope_dimension"])
        self.level1_admin.refresh_from_db()
        achievement_type = self._achievement_type()
        project = Project.objects.create(
            project_no="DC20260016",
            title="旧校级本学院成果项目",
            leader=self.student,
            status=Project.ProjectStatus.CLOSED,
            year=2026,
            batch=self.batch,
        )
        other_student = User.objects.create_user(
            username="legacy_school_achievement_leader",
            password="password123",
            role_fk=Role.objects.get(code="STUDENT"),
            real_name="旧校级其他学院成果负责人",
            employee_id="S90007",
            college="数学学院",
        )
        other_project = Project.objects.create(
            project_no="DC20260017",
            title="旧校级其他学院成果项目",
            leader=other_student,
            status=Project.ProjectStatus.CLOSED,
            year=2026,
            batch=self.batch,
        )
        ProjectAchievement.objects.create(
            project=project,
            achievement_type=achievement_type,
            title="旧校级本学院成果",
            description="D1",
            authors="A1",
        )
        ProjectAchievement.objects.create(
            project=other_project,
            achievement_type=achievement_type,
            title="旧校级其他学院成果",
            description="D2",
            authors="A2",
        )
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.get("/api/v1/projects/admin/achievements/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 2)

    def test_admin_project_list_tolerates_invalid_batch_filter(self):
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.get(
            "/api/v1/projects/admin/manage/",
            {"batch_id": "invalid"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["data"]["total"], 0)

    def test_admin_project_list_tolerates_invalid_dictionary_filters(self):
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.get(
            "/api/v1/projects/admin/manage/",
            {"level": "invalid", "category": "bad"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["data"]["total"], 0)

    def test_admin_cannot_direct_delete_non_draft_project(self):
        project = Project.objects.create(
            project_no="DC20260002",
            title="不可直接删除项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.delete(f"/api/v1/projects/admin/manage/{project.id}/")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["message"], "只有草稿项目可以直接删除")
        self.assertTrue(Project.objects.filter(id=project.id).exists())

    def test_admin_achievement_list_tolerates_invalid_year_filter(self):
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.get(
            "/api/v1/projects/admin/achievements/",
            {"year": "invalid"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 0)
        self.assertEqual(response.data["results"], [])

    def test_batch_status_rejects_invalid_project_ids(self):
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.post(
            "/api/v1/projects/admin/manage/batch-status/",
            {"project_ids": ["invalid"], "status": Project.ProjectStatus.CLOSED},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["message"], "项目ID列表不合法")

    def test_batch_status_requires_level1_admin(self):
        project = Project.objects.create(
            project_no="DC20260006",
            title="二级管理员批量状态测试项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        self.client.force_authenticate(user=self.level2_admin)

        response = self.client.post(
            "/api/v1/projects/admin/manage/batch-status/",
            {"project_ids": [project.id], "status": Project.ProjectStatus.CLOSED},
            format="json",
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.data["message"], "只有校级管理员可以批量变更项目状态")
        project.refresh_from_db()
        self.assertEqual(project.status, Project.ProjectStatus.IN_PROGRESS)

    def test_custom_school_admin_can_batch_update_project_status(self):
        project = Project.objects.create(
            project_no="DC20260014",
            title="自定义校级批量状态测试项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        self.client.force_authenticate(user=self.custom_school_admin)

        response = self.client.post(
            "/api/v1/projects/admin/manage/batch-status/",
            {"project_ids": [project.id], "status": Project.ProjectStatus.CLOSED},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["data"]["updated"], 1)
        project.refresh_from_db()
        self.assertEqual(project.status, Project.ProjectStatus.CLOSED)

    def test_batch_status_rejects_workflow_state_targets(self):
        project = Project.objects.create(
            project_no="DC20260007",
            title="批量状态流程目标测试项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.post(
            "/api/v1/projects/admin/manage/batch-status/",
            {"project_ids": [project.id], "status": Project.ProjectStatus.SUBMITTED},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["message"], "目标状态不合法")
        project.refresh_from_db()
        self.assertEqual(project.status, Project.ProjectStatus.IN_PROGRESS)

    def test_archive_closed_reports_service_failures(self):
        project = Project.objects.create(
            project_no="DC20260008",
            title="归档失败回传项目",
            leader=self.student,
            status=Project.ProjectStatus.CLOSED,
            year=2026,
            batch=self.batch,
        )
        self.client.force_authenticate(user=self.level1_admin)

        with patch(
            "apps.projects.views.mixins.project_batch_mixin."
            "ArchiveService.archive_projects",
            return_value={
                "created": [],
                "failed": [
                    {
                        "project_id": project.id,
                        "project_no": project.project_no,
                        "error": "archive write failed",
                    }
                ],
                "success_count": 0,
                "failed_count": 1,
            },
        ):
            response = self.client.post("/api/v1/projects/admin/manage/archive-closed/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["message"], "归档部分完成")
        self.assertEqual(response.data["data"]["created"], 0)
        self.assertEqual(response.data["data"]["skipped"], 0)
        self.assertEqual(response.data["data"]["failed_count"], 1)
        self.assertEqual(
            response.data["data"]["failures"][0]["error"],
            "archive write failed",
        )

    def test_level2_admin_archive_list_is_college_scoped(self):
        own_project = Project.objects.create(
            project_no="DC20260015",
            title="本学院归档项目",
            leader=self.student,
            status=Project.ProjectStatus.CLOSED,
            year=2026,
            batch=self.batch,
        )
        other_student = User.objects.create_user(
            username="other_college_archive_leader",
            password="password123",
            role_fk=Role.objects.get(code="STUDENT"),
            real_name="其他学院归档负责人",
            employee_id="S90006",
            college="数学学院",
        )
        other_project = Project.objects.create(
            project_no="DC20260016",
            title="其他学院归档项目",
            leader=other_student,
            status=Project.ProjectStatus.CLOSED,
            year=2026,
            batch=self.batch,
        )
        own_archive = ProjectArchive.objects.create(project=own_project)
        other_archive = ProjectArchive.objects.create(project=other_project)
        self.client.force_authenticate(user=self.level2_admin)

        response = self.client.get("/api/v1/projects/admin/manage/archives/")

        self.assertEqual(response.status_code, 200)
        archive_ids = {item["id"] for item in response.data["data"]}
        self.assertIn(own_archive.id, archive_ids)
        self.assertNotIn(other_archive.id, archive_ids)

    def test_legacy_level1_admin_without_scope_sees_all_archives(self):
        self.level1_role.scope_dimension = None
        self.level1_role.save(update_fields=["scope_dimension"])
        self.level1_admin.refresh_from_db()
        own_project = Project.objects.create(
            project_no="DC20260018",
            title="旧校级本学院归档项目",
            leader=self.student,
            status=Project.ProjectStatus.CLOSED,
            year=2026,
            batch=self.batch,
        )
        other_student = User.objects.create_user(
            username="legacy_school_archive_leader",
            password="password123",
            role_fk=Role.objects.get(code="STUDENT"),
            real_name="旧校级其他学院归档负责人",
            employee_id="S90008",
            college="数学学院",
        )
        other_project = Project.objects.create(
            project_no="DC20260019",
            title="旧校级其他学院归档项目",
            leader=other_student,
            status=Project.ProjectStatus.CLOSED,
            year=2026,
            batch=self.batch,
        )
        own_archive = ProjectArchive.objects.create(project=own_project)
        other_archive = ProjectArchive.objects.create(project=other_project)
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.get("/api/v1/projects/admin/manage/archives/")

        self.assertEqual(response.status_code, 200)
        archive_ids = {item["id"] for item in response.data["data"]}
        self.assertIn(own_archive.id, archive_ids)
        self.assertIn(other_archive.id, archive_ids)

    def test_export_rejects_invalid_project_ids(self):
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.get(
            "/api/v1/projects/admin/manage/export/",
            {"ids": "1,invalid"},
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["message"], "项目ID列表无效")

    def test_export_project_data_handles_project_members(self):
        project = Project.objects.create(
            project_no="DC20260002",
            title="导出成员测试项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        ProjectMember.objects.create(
            project=project,
            user=self.student,
            role=ProjectMember.MemberRole.LEADER,
        )
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.get("/api/v1/projects/admin/manage/export/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_batch_download_attachments_streams_storage_backed_files(self):
        project = Project.objects.create(
            project_no="DC20260020",
            title="远程存储附件项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        project.proposal_file = StorageBackedFile()
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.get(
            "/api/v1/projects/admin/manage/batch-download/",
            {"ids": str(project.id)},
        )

        self.assertEqual(response.status_code, 200)
        with zipfile.ZipFile(BytesIO(response.content)) as zf:
            self.assertEqual(
                zf.namelist(),
                ["DC20260020_远程存储附件项目/申请书.pdf"],
            )
            self.assertEqual(
                zf.read("DC20260020_远程存储附件项目/申请书.pdf"),
                b"proposal-content",
            )

    def test_level2_admin_cannot_export_other_college_project_doc(self):
        other_student = User.objects.create_user(
            username="other_college_leader",
            password="password123",
            role_fk=Role.objects.get(code="STUDENT"),
            real_name="其他学院负责人",
            employee_id="S90002",
            college="数学学院",
        )
        project = Project.objects.create(
            project_no="DC20260004",
            title="其他学院申报书测试项目",
            leader=other_student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )

        with patch(
            "apps.projects.views.mixins.project_admin_export_documents_mixin."
            "DocumentService.generate_project_doc"
        ) as generate_doc:
            response = self.client.get(
                f"/api/v1/projects/admin/manage/{project.id}/export-doc/"
            )

        self.assertEqual(response.status_code, 404)
        generate_doc.assert_not_called()

    def test_batch_export_doc_filters_other_college_projects(self):
        allowed_project = Project.objects.create(
            project_no="DC20260005",
            title="本学院申报书测试项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        other_student = User.objects.create_user(
            username="other_college_batch_leader",
            password="password123",
            role_fk=Role.objects.get(code="STUDENT"),
            real_name="其他学院批量负责人",
            employee_id="S90003",
            college="数学学院",
        )
        blocked_project = Project.objects.create(
            project_no="DC20260006",
            title="批量越权申报书测试项目",
            leader=other_student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )

        def fake_doc(project_id):
            return (
                BytesIO(f"doc-{project_id}".encode()),
                f"../unsafe:{project_id}?.docx",
            )

        with patch(
            "apps.projects.views.mixins.project_admin_export_documents_mixin."
            "DocumentService.generate_project_doc",
            side_effect=fake_doc,
        ) as generate_doc:
            response = self.client.get(
                "/api/v1/projects/admin/manage/batch-export-doc/",
                {"ids": f"{allowed_project.id},{blocked_project.id}"},
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [call.args[0] for call in generate_doc.call_args_list],
            [allowed_project.id],
        )
        with zipfile.ZipFile(BytesIO(response.content)) as zf:
            self.assertEqual(zf.namelist(), [f"unsafe_{allowed_project.id}_.docx"])

    def test_legacy_history_import_rejects_invalid_batch_id(self):
        self.client.force_authenticate(user=self.level1_admin)
        uploaded = SimpleUploadedFile(
            "history.xlsx",
            b"not a spreadsheet",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            "/api/v1/projects/admin/manage/import-history/",
            {"batch_id": "invalid", "file": uploaded},
            format="multipart",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["message"], "批次参数不合法")

    def test_legacy_history_import_rejects_missing_batch_id(self):
        self.client.force_authenticate(user=self.level1_admin)
        uploaded = SimpleUploadedFile(
            "history.xlsx",
            b"not a spreadsheet",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            "/api/v1/projects/admin/manage/import-history/",
            {"batch_id": "999999", "file": uploaded},
            format="multipart",
        )

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.data["message"], "批次不存在")

    def test_legacy_history_import_rejects_active_target_batch(self):
        self.client.force_authenticate(user=self.level1_admin)
        uploaded = SimpleUploadedFile(
            "history.xlsx",
            b"not a spreadsheet",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            "/api/v1/projects/admin/manage/import-history/",
            {"batch_id": str(self.batch.id), "file": uploaded},
            format="multipart",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["message"], "历史项目只能导入归档批次")

    def test_legacy_history_import_reports_invalid_year(self):
        self.client.force_authenticate(user=self.level1_admin)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["项目编号", "项目名称", "负责人学号/工号", "负责人姓名", "学院", "项目年份"])
        sheet.append(["LEG20260001", "旧导入年度错误项目", "S90001", "项目负责人", "计算机学院", "invalid"])

        stream = BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "history.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            "/api/v1/projects/admin/manage/import-history/",
            {"file": uploaded},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["data"]["created"], 0)
        self.assertEqual(response.data["data"]["errors"], ["第2行项目年份必须为四位数字"])

    def test_legacy_history_import_uses_archived_history_batch(self):
        self.client.force_authenticate(user=self.level1_admin)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["项目编号", "项目名称", "负责人学号/工号", "负责人姓名", "学院", "项目年份"])
        sheet.append(["LEG20260002", "旧导入当前年历史项目", "S90001", "项目负责人", "计算机学院", "2026"])
        stream = BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "history.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            "/api/v1/projects/admin/manage/import-history/",
            {"file": uploaded},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["data"]["created"], 1)
        project = Project.objects.get(project_no="LEG20260002")
        self.assertEqual(project.batch.code, "HIST2026")
        self.assertEqual(project.batch.status, ProjectBatch.STATUS_ARCHIVED)
        self.assertNotEqual(project.batch_id, self.batch.id)

    def test_custom_school_admin_can_import_history_projects(self):
        self.client.force_authenticate(user=self.custom_school_admin)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(
            ["项目编号", "项目名称", "负责人学号/工号", "负责人姓名", "学院", "项目年份"]
        )
        sheet.append(
            [
                "LEG20260004",
                "自定义校级导入历史项目",
                "S90001",
                "项目负责人",
                "计算机学院",
                "2026",
            ]
        )
        stream = BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "history.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            "/api/v1/projects/admin/manage/import-history/",
            {"file": uploaded},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["data"]["created"], 1)
        project = Project.objects.get(project_no="LEG20260004")
        self.assertEqual(project.batch.code, "HIST2026")
        self.assertEqual(project.batch.status, ProjectBatch.STATUS_ARCHIVED)

    def test_legacy_history_import_rejects_live_project_no_collision(self):
        self.client.force_authenticate(user=self.level1_admin)
        project = Project.objects.create(
            project_no="LEG20260003",
            title="在研项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["项目编号", "项目名称", "负责人学号/工号", "负责人姓名", "学院", "项目年份"])
        sheet.append(["LEG20260003", "历史覆盖尝试", "S90001", "项目负责人", "计算机学院", "2026"])
        stream = BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "history.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            "/api/v1/projects/admin/manage/import-history/",
            {"file": uploaded},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["data"]["created"], 0)
        self.assertEqual(
            response.data["data"]["errors"],
            ["第2行项目编号已存在于非历史批次，不能覆盖"],
        )
        project.refresh_from_db()
        self.assertEqual(project.title, "在研项目")
        self.assertEqual(project.status, Project.ProjectStatus.IN_PROGRESS)
        self.assertEqual(project.batch_id, self.batch.id)

    def test_admin_project_update_rejects_invalid_people_payload(self):
        teacher = User.objects.create_user(
            username="existing_advisor",
            password="password123",
            role_fk=Role.objects.get(code="TEACHER"),
            real_name="既有导师",
            employee_id="T90001",
        )
        project = Project.objects.create(
            project_no="DC20260003",
            title="管理员编辑人员测试项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        ProjectAdvisor.objects.create(project=project, user=teacher)
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.patch(
            f"/api/v1/projects/admin/manage/{project.id}/",
            {"title": "不应保存的标题", "advisors": "[invalid"},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["message"], "人员数据格式错误")
        project.refresh_from_db()
        self.assertEqual(project.title, "管理员编辑人员测试项目")
        self.assertTrue(
            ProjectAdvisor.objects.filter(project=project, user=teacher).exists()
        )

    def test_level1_admin_update_ignores_workflow_owned_project_fields(self):
        project = Project.objects.create(
            project_no="DC20260004",
            title="管理员编辑状态测试项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        other_batch = ProjectBatch.objects.create(
            name="2027",
            year=2027,
            code="B2027",
            status=ProjectBatch.STATUS_ACTIVE,
            is_active=True,
            is_current=False,
        )
        self.client.force_authenticate(user=self.level1_admin)

        response = self.client.patch(
            f"/api/v1/projects/admin/manage/{project.id}/",
            {
                "title": "允许修改的标题",
                "status": Project.ProjectStatus.CLOSED,
                "batch": other_batch.id,
                "publish_status": Project.PublishStatus.PUBLISHED,
                "final_budget": "9999.00",
                "is_deleted": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        project.refresh_from_db()
        self.assertEqual(project.title, "允许修改的标题")
        self.assertEqual(project.status, Project.ProjectStatus.IN_PROGRESS)
        self.assertEqual(project.batch_id, self.batch.id)
        self.assertEqual(project.publish_status, Project.PublishStatus.NOT_READY)
        self.assertIsNone(project.final_budget)
        self.assertFalse(project.is_deleted)

    def test_custom_school_admin_update_uses_school_forbidden_fields(self):
        project = Project.objects.create(
            project_no="DC20260011",
            title="自定义校级管理员编辑项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        self.client.force_authenticate(user=self.custom_school_admin)

        response = self.client.patch(
            f"/api/v1/projects/admin/manage/{project.id}/",
            {
                "title": "自定义校级允许修改的标题",
                "approved_budget": "1234.00",
                "final_budget": "9999.00",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        project.refresh_from_db()
        self.assertEqual(project.title, "自定义校级允许修改的标题")
        self.assertEqual(project.approved_budget, Decimal("1234.00"))
        self.assertIsNone(project.final_budget)

    def test_level2_admin_update_ignores_approved_budget(self):
        project = Project.objects.create(
            project_no="DC20260005",
            title="二级管理员经费测试项目",
            leader=self.student,
            status=Project.ProjectStatus.IN_PROGRESS,
            year=2026,
            batch=self.batch,
        )
        self.client.force_authenticate(user=self.level2_admin)

        response = self.client.patch(
            f"/api/v1/projects/admin/manage/{project.id}/",
            {"title": "二级管理员允许修改的标题", "approved_budget": "9999.00"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        project.refresh_from_db()
        self.assertEqual(project.title, "二级管理员允许修改的标题")
        self.assertIsNone(project.approved_budget)
