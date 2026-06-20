import os
import hashlib
import uuid
import logging
from pathlib import Path, PurePosixPath, PureWindowsPath
from decimal import Decimal, InvalidOperation

import openpyxl  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]
from django.conf import settings
from django.core.files import File
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from apps.dictionaries.models import DictionaryItem, DictionaryType
from apps.projects.models import Project
from apps.system_settings.models import ProjectBatch
from apps.users.models import User
from apps.users.services import UserService
from apps.users.models import Role

from .models import AsyncTaskRecord, OperationLog

logger = logging.getLogger(__name__)


IMPORT_KINDS = {
    "students": {
        "title": "学生数据导入",
        "role": User.UserRole.STUDENT,
        "headers": ["学号", "姓名", "单位名称", "专业名称", "当前年级", "班级", "性别"],
        "required_headers": ["学号", "姓名"],
    },
    "teachers": {
        "title": "教师数据导入",
        "role": User.UserRole.TEACHER,
        "headers": ["Tno", "TName", "Cname", "Ranks"],
        "required_headers": ["Tno", "TName"],
    },
    "level2_admins": {
        "title": "二级管理员导入",
        "role": User.UserRole.LEVEL2_ADMIN,
        "headers": ["工号", "姓名", "学院", "部门", "手机号", "邮箱"],
        "required_headers": ["工号", "姓名"],
    },
    "colleges": {
        "title": "学院数据导入",
        "role": None,
        "headers": ["代码", "名称", "排序"],
        "required_headers": ["代码", "名称"],
    },
    "majors": {
        "title": "专业数据导入",
        "role": None,
        "headers": ["代码", "名称", "学院", "排序"],
        "required_headers": ["代码", "名称"],
    },
    "history_projects": {
        "title": "历史项目导入",
        "role": None,
        "headers": ["项目编号", "项目名称", "负责人", "学院", "级别", "经费", "年度"],
        "required_headers": ["项目编号", "项目名称"],
    },
    "dictionaries": {
        "title": "系统字典导入",
        "role": None,
        "headers": ["字典类型", "类型名称", "代码", "名称", "排序"],
        "required_headers": ["字典类型", "代码", "名称"],
    },
}

ALLOWED_IMPORT_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
MAX_IMPORT_UPLOAD_SIZE = 10 * 1024 * 1024


def _safe_task_error_message(exc):
    if isinstance(exc, ValueError):
        return str(exc)
    return "导入失败，请检查文件内容后重试"


def _cleanup_local_file(file_path):
    if not file_path:
        return
    try:
        Path(file_path).unlink(missing_ok=True)
    except OSError:
        logger.warning("Failed to remove import source file: %s", file_path)


def _open_workbook(file_path):
    try:
        return openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except (InvalidFileException, OSError, KeyError, ValueError) as exc:
        raise ValueError("导入文件格式错误，请上传有效的 Excel 文件") from exc


class OperationLogService:
    @staticmethod
    def log(
        *,
        operator=None,
        module,
        action,
        target_type="",
        target_id="",
        target_name="",
        status=OperationLog.LogStatus.SUCCESS,
        detail=None,
        request=None,
    ):
        ip_address = None
        if request:
            forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR", "")
            ip_address = (
                forwarded_for.split(",")[0].strip()
                if forwarded_for
                else request.META.get("REMOTE_ADDR")
            )
        return OperationLog.objects.create(
            operator=operator if getattr(operator, "is_authenticated", True) else None,
            module=module,
            action=action,
            target_type=target_type,
            target_id=str(target_id) if target_id not in (None, "") else "",
            target_name=target_name or "",
            status=status,
            detail=detail or {},
            ip_address=ip_address,
        )


class DataCenterService:
    @staticmethod
    def get_import_kinds():
        return [
            {"kind": kind, "title": config["title"], "headers": config["headers"]}
            for kind, config in IMPORT_KINDS.items()
        ]

    @staticmethod
    def build_template(kind):
        config = IMPORT_KINDS.get(kind)
        if not config:
            raise ValueError("不支持的导入类型")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = config["title"][:31]
        sheet.append(config["headers"])
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        return workbook

    @staticmethod
    def save_upload(uploaded_file):
        raw_name = str(uploaded_file.name or "")
        original_name = PureWindowsPath(PurePosixPath(raw_name).name).name
        suffix = Path(original_name).suffix.lower()
        if suffix not in ALLOWED_IMPORT_EXTENSIONS:
            raise ValueError("仅支持 xlsx/xlsm/xltx/xltm 文件")

        file_size = getattr(uploaded_file, "size", 0) or 0
        if file_size > MAX_IMPORT_UPLOAD_SIZE:
            raise ValueError("导入文件不能超过 10MB")

        upload_dir = Path(settings.LOCAL_DATA_DIR) / "imports"
        upload_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{timezone.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex}{suffix}"
        path = upload_dir / filename
        with path.open("wb") as output:
            for chunk in uploaded_file.chunks():
                output.write(chunk)
        return str(path)

    @staticmethod
    def preview_file(kind, file_path):
        config = IMPORT_KINDS.get(kind)
        if not config:
            raise ValueError("不支持的导入类型")

        workbook = _open_workbook(file_path)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        header = [str(cell).strip() if cell is not None else "" for cell in rows[0]] if rows else []
        required_headers = config.get("required_headers", config["headers"][:2])
        missing = [name for name in required_headers if name not in header]
        return {
            "headers": header,
            "total_rows": max(len(rows) - 1, 0),
            "missing_headers": missing,
            "valid": not missing,
        }

    @staticmethod
    def _load_rows(file_path):
        workbook = _open_workbook(file_path)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if not rows:
            return [], []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        return headers, rows[1:]

    @staticmethod
    def _row_dict(headers, row):
        data = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row[index] if index < len(row) else ""
            data[header] = str(value).strip() if value is not None else ""
        return data

    @staticmethod
    def _decimal(value):
        if value in (None, ""):
            return Decimal("0")
        try:
            amount = Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError) as exc:
            raise ValueError("必须为数字") from exc
        if not amount.is_finite():
            raise ValueError("必须为有效数字")
        if amount < 0:
            raise ValueError("不能为负数")
        return amount

    @staticmethod
    def _integer(value, default=0, minimum=None):
        if value in (None, ""):
            return default
        try:
            parsed = int(str(value).strip())
        except (TypeError, ValueError) as exc:
            raise ValueError("必须为整数") from exc
        if minimum is not None and parsed < minimum:
            raise ValueError(f"不能小于 {minimum}")
        return parsed

    @staticmethod
    def _history_year(value, project_no):
        year_text = str(value).strip() if value not in (None, "") else ""
        if not year_text:
            year_text = str(project_no).strip()[:4]
        if len(year_text) != 4 or not year_text.isdigit():
            raise ValueError("必须为四位数字")
        year = int(year_text)
        max_year = timezone.now().year + 1
        if year < 1900 or year > max_year:
            raise ValueError(f"必须在 1900 至 {max_year} 年之间")
        return year

    @staticmethod
    def _resolve_project_level(value):
        if not value:
            return None
        return DictionaryItem.objects.filter(
            dict_type__code="project_level"
        ).filter(Q(value=value) | Q(label=value)).first()

    @staticmethod
    def _history_batch(year):
        batch, _ = ProjectBatch.objects.get_or_create(
            code=f"HIST{year}",
            defaults={
                "name": f"{year} 年历史项目",
                "year": year,
                "status": ProjectBatch.STATUS_ARCHIVED,
                "is_current": False,
                "is_active": True,
            },
        )
        if batch.status != ProjectBatch.STATUS_ARCHIVED:
            batch.status = ProjectBatch.STATUS_ARCHIVED
            batch.is_current = False
            batch.is_active = True
            batch.save(update_fields=["status", "is_current", "is_active"])
        return batch

    @staticmethod
    def _history_leader(project_no, real_name, college):
        role = Role.objects.filter(code=User.UserRole.STUDENT).first()
        if not role:
            raise ValueError("学生角色不存在")
        digest = hashlib.sha1((project_no or real_name).encode("utf-8")).hexdigest()
        employee_id = f"H{digest[:12].upper()}"
        user, created = User.objects.get_or_create(
            employee_id=employee_id,
            defaults={
                "username": employee_id,
                "role_fk": role,
                "real_name": real_name or "历史项目负责人",
                "college": college,
                "is_active": True,
            },
        )
        changed_fields = []
        if created:
            user.set_unusable_password()
            changed_fields.append("password")
        if real_name and user.real_name != real_name:
            user.real_name = real_name
            changed_fields.append("real_name")
        if college and user.college != college:
            user.college = college
            changed_fields.append("college")
        if user.role_fk_id != role.id:
            user.role_fk = role
            changed_fields.append("role_fk")
        if changed_fields:
            user.save(update_fields=changed_fields)
        return user

    @staticmethod
    @transaction.atomic
    def _import_history_projects(file_path):
        headers, rows = DataCenterService._load_rows(file_path)
        created = 0
        updated = 0
        errors = []
        for row_index, row in enumerate(rows, start=2):
            row_data = DataCenterService._row_dict(headers, row)
            if not any(row_data.values()):
                continue
            project_no = row_data.get("项目编号", "")
            title = row_data.get("项目名称", "")
            if not project_no or not title:
                errors.append(f"Row {row_index}: 项目编号和项目名称不能为空")
                continue
            try:
                year = DataCenterService._history_year(
                    row_data.get("年度", ""), project_no
                )
            except ValueError as exc:
                errors.append(f"Row {row_index}: 年度{exc}")
                continue
            college = row_data.get("学院", "")
            leader = DataCenterService._history_leader(
                project_no, row_data.get("负责人", ""), college
            )
            level = DataCenterService._resolve_project_level(row_data.get("级别", ""))
            batch = DataCenterService._history_batch(year)
            try:
                budget = DataCenterService._decimal(row_data.get("经费", ""))
            except ValueError as exc:
                errors.append(f"Row {row_index}: 经费{exc}")
                continue
            existing_project = Project.objects.filter(project_no=project_no).first()
            if existing_project and existing_project.batch_id != batch.id:
                errors.append(
                    f"Row {row_index}: 项目编号已存在于非历史批次，不能覆盖"
                )
                continue
            project, was_created = Project.objects.update_or_create(
                project_no=project_no,
                defaults={
                    "title": title,
                    "leader": leader,
                    "year": year,
                    "batch": batch,
                    "level": level,
                    "budget": budget,
                    "approved_budget": budget,
                    "final_level": level,
                    "final_budget": budget,
                    "status": Project.ProjectStatus.CLOSED,
                    "publish_status": Project.PublishStatus.PUBLISHED,
                    "published_at": timezone.now(),
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1
        return {"created": created, "updated": updated, "errors": errors}

    @staticmethod
    @transaction.atomic
    def _import_dictionaries(file_path):
        headers, rows = DataCenterService._load_rows(file_path)
        created = 0
        updated = 0
        errors = []
        for row_index, row in enumerate(rows, start=2):
            row_data = DataCenterService._row_dict(headers, row)
            if not any(row_data.values()):
                continue
            type_code = row_data.get("字典类型", "")
            value = row_data.get("代码", "")
            label = row_data.get("名称", "")
            if not type_code or not value or not label:
                errors.append(f"Row {row_index}: 字典类型、代码、名称不能为空")
                continue
            try:
                sort_order = DataCenterService._integer(
                    row_data.get("排序", ""), default=0, minimum=0
                )
            except ValueError as exc:
                errors.append(f"Row {row_index}: 排序{exc}")
                continue
            dict_type, _ = DictionaryType.objects.get_or_create(
                code=type_code,
                defaults={
                    "name": row_data.get("类型名称", "") or type_code,
                    "is_active": True,
                },
            )
            item, was_created = DictionaryItem.objects.update_or_create(
                dict_type=dict_type,
                value=value,
                defaults={
                    "label": label,
                    "sort_order": sort_order,
                    "is_active": True,
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1
        return {"created": created, "updated": updated, "errors": errors}

    @staticmethod
    @transaction.atomic
    def _import_simple_dictionary(file_path, type_code, type_name, extra_builder=None):
        headers, rows = DataCenterService._load_rows(file_path)
        dict_type, _ = DictionaryType.objects.get_or_create(
            code=type_code,
            defaults={"name": type_name, "is_active": True},
        )
        created = 0
        updated = 0
        errors = []
        for row_index, row in enumerate(rows, start=2):
            row_data = DataCenterService._row_dict(headers, row)
            if not any(row_data.values()):
                continue
            value = row_data.get("代码", "")
            label = row_data.get("名称", "")
            if not value or not label:
                errors.append(f"Row {row_index}: 代码和名称不能为空")
                continue
            try:
                sort_order = DataCenterService._integer(
                    row_data.get("排序", ""), default=0, minimum=0
                )
            except ValueError as exc:
                errors.append(f"Row {row_index}: 排序{exc}")
                continue
            item, was_created = DictionaryItem.objects.update_or_create(
                dict_type=dict_type,
                value=value,
                defaults={
                    "label": label,
                    "sort_order": sort_order,
                    "is_active": True,
                    "extra_data": extra_builder(row_data) if extra_builder else {},
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1
        return {"created": created, "updated": updated, "errors": errors}

    @staticmethod
    def create_import_task(user, kind, uploaded_file):
        config = IMPORT_KINDS.get(kind)
        if not config:
            raise ValueError("不支持的导入类型")
        file_path = DataCenterService.save_upload(uploaded_file)
        try:
            preview = DataCenterService.preview_file(kind, file_path)
        except Exception:
            _cleanup_local_file(file_path)
            raise
        task = AsyncTaskRecord.objects.create(
            task_type=AsyncTaskRecord.TaskType.IMPORT,
            title=config["title"],
            status=AsyncTaskRecord.TaskStatus.PENDING,
            progress=0,
            message="任务已创建",
            payload={"kind": kind, "file_path": file_path, "preview": preview},
            created_by=user,
        )
        return task

    @staticmethod
    def run_import_task(task_id):
        started_at = timezone.now()
        claimed = AsyncTaskRecord.objects.filter(
            id=task_id,
            status=AsyncTaskRecord.TaskStatus.PENDING,
        ).update(
            status=AsyncTaskRecord.TaskStatus.RUNNING,
            progress=20,
            started_at=started_at,
            message="正在执行导入",
        )
        task = AsyncTaskRecord.objects.get(id=task_id)
        if not claimed:
            return task

        file_path = task.payload.get("file_path")
        try:
            kind = task.payload.get("kind")
            config = IMPORT_KINDS.get(kind)
            if not config or not file_path or not os.path.exists(file_path):
                raise ValueError("任务文件不存在或类型无效")
            if task.payload.get("preview", {}).get("valid") is False:
                raise ValueError("模板表头不完整")

            if config["role"]:
                with open(file_path, "rb") as source:
                    upload = File(source, name=os.path.basename(file_path))
                    result = UserService().import_users(upload, default_role=config["role"])
            elif kind == "history_projects":
                result = DataCenterService._import_history_projects(file_path)
            elif kind == "dictionaries":
                result = DataCenterService._import_dictionaries(file_path)
            elif kind == "colleges":
                result = DataCenterService._import_simple_dictionary(
                    file_path, "college", "学院"
                )
            elif kind == "majors":
                result = DataCenterService._import_simple_dictionary(
                    file_path,
                    "major_category",
                    "专业大类",
                    lambda row: {"college": row.get("学院", "")},
                )
            else:
                result = {
                    "created": 0,
                    "errors": ["该类型已完成模板校验，正式写入将在专项模块中处理"],
                }

            task.status = AsyncTaskRecord.TaskStatus.SUCCESS
            task.progress = 100
            task.message = "导入完成"
            task.result = result
            task.completed_at = timezone.now()
            task.save(
                update_fields=[
                    "status",
                    "progress",
                    "message",
                    "result",
                    "completed_at",
                ]
            )
            OperationLogService.log(
                operator=task.created_by,
                module="数据中心",
                action="执行导入",
                target_type="AsyncTaskRecord",
                target_id=task.id,
                target_name=task.title,
                detail={"kind": kind, "result": result},
            )
        except Exception as exc:
            message = _safe_task_error_message(exc)
            logger.exception(
                "Failed to run import task %s for kind %s",
                task.id,
                task.payload.get("kind"),
            )
            task.status = AsyncTaskRecord.TaskStatus.FAILED
            task.progress = 100
            task.message = message
            task.result = {"errors": [message]}
            task.completed_at = timezone.now()
            task.save(
                update_fields=[
                    "status",
                    "progress",
                    "message",
                    "result",
                    "completed_at",
                ]
            )
            OperationLogService.log(
                operator=task.created_by,
                module="数据中心",
                action="执行导入",
                target_type="AsyncTaskRecord",
                target_id=task.id,
                target_name=task.title,
                status=OperationLog.LogStatus.FAILED,
                detail={"error": message, "kind": task.payload.get("kind")},
            )
        finally:
            _cleanup_local_file(file_path)
        return task

    @staticmethod
    def create_completed_file_task(user, title, task_type, filename, content, result=None):
        task = AsyncTaskRecord.objects.create(
            task_type=task_type,
            title=title,
            status=AsyncTaskRecord.TaskStatus.SUCCESS,
            progress=100,
            message="生成完成",
            result=result or {},
            created_by=user,
            started_at=timezone.now(),
            completed_at=timezone.now(),
        )
        task.result_file.save(filename, ContentFile(content), save=True)
        OperationLogService.log(
            operator=user,
            module="任务中心",
            action="生成文件任务",
            target_type="AsyncTaskRecord",
            target_id=task.id,
            target_name=title,
            detail={"filename": filename, "task_type": task_type},
        )
        return task
