from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl  # type: ignore[import-untyped]
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from rest_framework.test import APIClient

from apps.dictionaries.models import DictionaryItem, DictionaryType
from apps.operations.models import AsyncTaskRecord, OperationLog
from apps.operations.services import DataCenterService
from apps.projects.models import Project
from apps.system_settings.models import ProjectBatch
from apps.users.models import Role


User = get_user_model()


class DataCenterServiceTestCase(TestCase):
    def setUp(self):
        self.temp_dir = TemporaryDirectory()
        self.settings_override = override_settings(
            LOCAL_DATA_DIR=self.temp_dir.name,
            MEDIA_ROOT=self.temp_dir.name,
            DEFAULT_USER_PASSWORD="import-password-123",
        )
        self.settings_override.enable()

        self.student_role = self._role("STUDENT", "学生")
        self.admin_role = self._role("LEVEL1_ADMIN", "校级管理员", "SCHOOL")
        self.custom_school_role = self._role(
            "SCHOOL_DATA_CENTER_ADMIN",
            "校级数据中心管理员",
            "SCHOOL",
        )
        self.admin = User.objects.create_user(
            username="data_admin",
            password="123456",
            role_fk=self.admin_role,
            real_name="数据管理员",
            employee_id="DATA0001",
        )
        self.custom_school_admin = User.objects.create_user(
            username="custom_data_admin",
            password="123456",
            role_fk=self.custom_school_role,
            real_name="校级数据中心管理员",
            employee_id="DATA0002",
        )
        self.other_admin = User.objects.create_user(
            username="other_data_admin",
            password="123456",
            role_fk=self.admin_role,
            real_name="其他数据管理员",
            employee_id="DATA0003",
        )
        self.level_type, _ = DictionaryType.objects.get_or_create(
            code="project_level", defaults={"name": "项目级别"}
        )
        self.school_level, _ = DictionaryItem.objects.get_or_create(
            dict_type=self.level_type,
            value="SCHOOL",
            defaults={"label": "校级"},
        )
        self.client = APIClient()

    def tearDown(self):
        self.settings_override.disable()
        self.temp_dir.cleanup()

    def _role(self, code, name, scope_dimension=None):
        role, _ = Role.objects.get_or_create(code=code, defaults={"name": name})
        role.name = name
        role.scope_dimension = scope_dimension
        role.is_active = True
        role.save(update_fields=["name", "scope_dimension", "is_active"])
        return role

    def _uploaded_students_file(self):
        workbook = DataCenterService.build_template("students")
        sheet = workbook.active
        sheet.append(["S20260001", "学生导入测试", "计算机学院", "软件工程", "2026", "1班", "男"])
        stream = BytesIO()
        workbook.save(stream)
        return SimpleUploadedFile(
            "students.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_build_template_and_run_import_task(self):
        workbook = DataCenterService.build_template("students")
        self.assertEqual(workbook.active.title, "学生数据导入")
        self.assertEqual(workbook.active["A1"].value, "学号")

        task = DataCenterService.create_import_task(
            self.admin,
            "students",
            self._uploaded_students_file(),
        )

        self.assertEqual(task.status, AsyncTaskRecord.TaskStatus.PENDING)
        self.assertTrue(task.payload["preview"]["valid"])
        self.assertEqual(task.payload["preview"]["total_rows"], 1)

        finished = DataCenterService.run_import_task(task.id)

        self.assertEqual(finished.status, AsyncTaskRecord.TaskStatus.SUCCESS)
        self.assertEqual(finished.progress, 100)
        self.assertEqual(finished.result["created"], 1)
        self.assertTrue(
            OperationLog.objects.filter(
                module="数据中心",
                action="执行导入",
                target_id=str(finished.id),
            ).exists()
        )
        imported = User.objects.get(employee_id="S20260001")
        self.assertEqual(imported.role_fk, self.student_role)
        self.assertEqual(imported.real_name, "学生导入测试")
        self.assertEqual(imported.college, "计算机学院")

    def test_custom_school_admin_can_access_data_center_kinds(self):
        self.client.force_authenticate(user=self.custom_school_admin)

        response = self.client.get("/api/v1/operations/tasks/data-center/kinds/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["code"], 200)
        kind_codes = {item["kind"] for item in response.data["data"]}
        self.assertIn("students", kind_codes)

    def test_report_snapshot_endpoint_is_removed(self):
        self.client.force_authenticate(user=self.custom_school_admin)

        response = self.client.post(
            "/api/v1/operations/tasks/report/snapshot/",
            {"title": "removed", "result": {"total": 1}},
            format="json",
        )

        self.assertEqual(response.status_code, 404)

    def test_statistics_report_endpoints_are_removed(self):
        self.client.force_authenticate(user=self.custom_school_admin)

        report_response = self.client.get(
            "/api/v1/projects/admin/manage/statistics-report/"
        )
        task_response = self.client.post(
            "/api/v1/projects/admin/manage/statistics-report/task/",
            {},
            format="json",
        )

        self.assertEqual(report_response.status_code, 404)
        self.assertEqual(task_response.status_code, 404)

    def test_custom_school_admin_sees_all_tasks_and_logs(self):
        own_task = AsyncTaskRecord.objects.create(
            task_type=AsyncTaskRecord.TaskType.EXPORT,
            title="own",
            created_by=self.custom_school_admin,
        )
        other_task = AsyncTaskRecord.objects.create(
            task_type=AsyncTaskRecord.TaskType.EXPORT,
            title="other",
            created_by=self.other_admin,
        )
        own_log = OperationLog.objects.create(
            operator=self.custom_school_admin,
            module="测试",
            action="自己的日志",
        )
        other_log = OperationLog.objects.create(
            operator=self.other_admin,
            module="测试",
            action="其他日志",
        )
        self.client.force_authenticate(user=self.custom_school_admin)

        tasks_response = self.client.get("/api/v1/operations/tasks/")
        logs_response = self.client.get("/api/v1/operations/logs/")

        self.assertEqual(tasks_response.status_code, 200)
        self.assertEqual(logs_response.status_code, 200)
        task_ids = {item["id"] for item in tasks_response.data["results"]}
        log_ids = {item["id"] for item in logs_response.data["results"]}
        self.assertIn(own_task.id, task_ids)
        self.assertIn(other_task.id, task_ids)
        self.assertIn(own_log.id, log_ids)
        self.assertIn(other_log.id, log_ids)

    def test_legacy_level1_admin_retains_school_operations_access(self):
        self.admin_role.scope_dimension = None
        self.admin_role.save(update_fields=["scope_dimension"])
        self.admin.refresh_from_db()

        own_task = AsyncTaskRecord.objects.create(
            task_type=AsyncTaskRecord.TaskType.EXPORT,
            title="legacy-own",
            created_by=self.admin,
        )
        other_task = AsyncTaskRecord.objects.create(
            task_type=AsyncTaskRecord.TaskType.EXPORT,
            title="legacy-other",
            created_by=self.other_admin,
        )
        own_log = OperationLog.objects.create(
            operator=self.admin,
            module="测试",
            action="旧校级自己的日志",
        )
        other_log = OperationLog.objects.create(
            operator=self.other_admin,
            module="测试",
            action="旧校级其他日志",
        )
        self.client.force_authenticate(user=self.admin)

        kinds_response = self.client.get("/api/v1/operations/tasks/data-center/kinds/")
        tasks_response = self.client.get("/api/v1/operations/tasks/")
        logs_response = self.client.get("/api/v1/operations/logs/")

        self.assertEqual(kinds_response.status_code, 200)
        self.assertEqual(kinds_response.data["code"], 200)
        kind_codes = {item["kind"] for item in kinds_response.data["data"]}
        self.assertIn("students", kind_codes)
        self.assertEqual(tasks_response.status_code, 200)
        self.assertEqual(logs_response.status_code, 200)
        task_ids = {item["id"] for item in tasks_response.data["results"]}
        log_ids = {item["id"] for item in logs_response.data["results"]}
        self.assertIn(own_task.id, task_ids)
        self.assertIn(other_task.id, task_ids)
        self.assertIn(own_log.id, log_ids)
        self.assertIn(other_log.id, log_ids)

    def test_run_import_task_does_not_reprocess_finished_task(self):
        task = DataCenterService.create_import_task(
            self.admin,
            "students",
            self._uploaded_students_file(),
        )
        finished = DataCenterService.run_import_task(task.id)

        rerun = DataCenterService.run_import_task(task.id)

        self.assertEqual(rerun.status, AsyncTaskRecord.TaskStatus.SUCCESS)
        self.assertEqual(rerun.result, finished.result)
        self.assertEqual(User.objects.filter(employee_id="S20260001").count(), 1)

    def test_preview_reports_missing_required_headers(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["错误列", "姓名"])
        stream = BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "invalid.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        file_path = DataCenterService.save_upload(uploaded)

        preview = DataCenterService.preview_file("students", file_path)

        self.assertFalse(preview["valid"])
        self.assertEqual(preview["missing_headers"], ["学号"])

    def test_preview_uses_kind_specific_required_headers(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["字典类型", "类型名称", "排序"])
        stream = BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "invalid-dictionary.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        file_path = DataCenterService.save_upload(uploaded)

        preview = DataCenterService.preview_file("dictionaries", file_path)

        self.assertFalse(preview["valid"])
        self.assertEqual(preview["missing_headers"], ["代码", "名称"])

    def test_import_task_rejects_kind_specific_missing_headers(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["字典类型", "类型名称", "排序"])
        sheet.append(["demo_type", "演示字典", "10"])
        stream = BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "invalid-dictionary-task.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        task = DataCenterService.create_import_task(self.admin, "dictionaries", uploaded)
        finished = DataCenterService.run_import_task(task.id)

        self.assertEqual(finished.status, AsyncTaskRecord.TaskStatus.FAILED)
        self.assertEqual(finished.message, "模板表头不完整")
        self.assertEqual(finished.result, {"errors": ["模板表头不完整"]})
        self.assertFalse(DictionaryItem.objects.filter(dict_type__code="demo_type").exists())

    def test_save_upload_sanitizes_filename(self):
        uploaded = SimpleUploadedFile(
            "../unsafe.xlsx",
            self._uploaded_students_file().read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        file_path = DataCenterService.save_upload(uploaded)
        saved_path = Path(file_path)

        self.assertTrue(saved_path.name.endswith(".xlsx"))
        self.assertNotIn("unsafe", saved_path.name)
        self.assertTrue(
            saved_path.resolve().is_relative_to(Path(self.temp_dir.name).resolve())
        )

    def test_save_upload_rejects_unsupported_file_type(self):
        uploaded = SimpleUploadedFile(
            "students.txt",
            b"not a spreadsheet",
            content_type="text/plain",
        )

        with self.assertRaisesMessage(ValueError, "仅支持 xlsx/xlsm/xltx/xltm 文件"):
            DataCenterService.save_upload(uploaded)

    def test_save_upload_rejects_large_file(self):
        uploaded = SimpleUploadedFile(
            "students.xlsx",
            b"0" * (10 * 1024 * 1024 + 1),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with self.assertRaisesMessage(ValueError, "导入文件不能超过 10MB"):
            DataCenterService.save_upload(uploaded)

    def test_import_history_projects_into_archived_batch(self):
        workbook = DataCenterService.build_template("history_projects")
        sheet = workbook.active
        sheet.append(["H20240001", "历史项目导入测试", "历史负责人", "计算机学院", "校级", "1200", "2024"])
        stream = BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "history.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        task = DataCenterService.create_import_task(
            self.admin,
            "history_projects",
            uploaded,
        )
        finished = DataCenterService.run_import_task(task.id)

        self.assertEqual(finished.status, AsyncTaskRecord.TaskStatus.SUCCESS)
        self.assertEqual(finished.result["created"], 1)
        project = Project.objects.get(project_no="H20240001")
        self.assertEqual(project.status, Project.ProjectStatus.CLOSED)
        self.assertEqual(project.publish_status, Project.PublishStatus.PUBLISHED)
        self.assertEqual(project.batch.status, ProjectBatch.STATUS_ARCHIVED)
        self.assertEqual(project.leader.real_name, "历史负责人")
        self.assertEqual(project.final_budget, 1200)

    def test_import_history_projects_uses_archived_batch_for_current_year(self):
        current_batch = ProjectBatch.objects.create(
            name="2026 当前批次",
            year=2026,
            code="CUR2026",
            status=ProjectBatch.STATUS_ACTIVE,
            is_current=True,
            is_active=True,
        )
        workbook = DataCenterService.build_template("history_projects")
        sheet = workbook.active
        sheet.append(["H20260001", "当前年度历史项目", "历史负责人", "计算机学院", "校级", "1200", "2026"])
        stream = BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "history-current-year.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        task = DataCenterService.create_import_task(
            self.admin,
            "history_projects",
            uploaded,
        )
        finished = DataCenterService.run_import_task(task.id)

        self.assertEqual(finished.status, AsyncTaskRecord.TaskStatus.SUCCESS)
        project = Project.objects.get(project_no="H20260001")
        self.assertNotEqual(project.batch_id, current_batch.id)
        self.assertEqual(project.batch.code, "HIST2026")
        self.assertEqual(project.batch.status, ProjectBatch.STATUS_ARCHIVED)

    def test_import_history_projects_rejects_live_project_no_collision(self):
        current_batch = ProjectBatch.objects.create(
            name="2026 当前批次",
            year=2026,
            code="LIVE2026",
            status=ProjectBatch.STATUS_ACTIVE,
            is_current=True,
            is_active=True,
        )
        live_project = Project.objects.create(
            project_no="DC20260001",
            title="在研项目",
            leader=self.admin,
            year=2026,
            batch=current_batch,
            status=Project.ProjectStatus.IN_PROGRESS,
        )
        workbook = DataCenterService.build_template("history_projects")
        sheet = workbook.active
        sheet.append(["DC20260001", "历史覆盖尝试", "历史负责人", "计算机学院", "校级", "1200", "2026"])
        stream = BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "history-live-collision.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        task = DataCenterService.create_import_task(
            self.admin,
            "history_projects",
            uploaded,
        )
        finished = DataCenterService.run_import_task(task.id)

        self.assertEqual(finished.status, AsyncTaskRecord.TaskStatus.SUCCESS)
        self.assertEqual(finished.result["created"], 0)
        self.assertEqual(
            finished.result["errors"],
            ["Row 2: 项目编号已存在于非历史批次，不能覆盖"],
        )
        live_project.refresh_from_db()
        self.assertEqual(live_project.title, "在研项目")
        self.assertEqual(live_project.status, Project.ProjectStatus.IN_PROGRESS)
        self.assertEqual(live_project.batch_id, current_batch.id)

    def test_import_history_projects_rejects_invalid_year(self):
        workbook = DataCenterService.build_template("history_projects")
        sheet = workbook.active
        sheet.append(
            ["H20240002", "年度错误项目", "历史负责人", "计算机学院", "校级", "1200", "invalid"]
        )
        stream = BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "history-invalid-year.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        task = DataCenterService.create_import_task(
            self.admin,
            "history_projects",
            uploaded,
        )
        finished = DataCenterService.run_import_task(task.id)

        self.assertEqual(finished.status, AsyncTaskRecord.TaskStatus.SUCCESS)
        self.assertEqual(finished.result["created"], 0)
        self.assertEqual(finished.result["errors"], ["Row 2: 年度必须为四位数字"])
        self.assertFalse(Project.objects.filter(project_no="H20240002").exists())

    def test_import_history_projects_rejects_negative_budget(self):
        workbook = DataCenterService.build_template("history_projects")
        sheet = workbook.active
        sheet.append(["H20240003", "经费错误项目", "历史负责人", "计算机学院", "校级", "-1", "2024"])
        stream = BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "history-negative-budget.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        task = DataCenterService.create_import_task(
            self.admin,
            "history_projects",
            uploaded,
        )
        finished = DataCenterService.run_import_task(task.id)

        self.assertEqual(finished.status, AsyncTaskRecord.TaskStatus.SUCCESS)
        self.assertEqual(finished.result["created"], 0)
        self.assertEqual(finished.result["errors"], ["Row 2: 经费不能为负数"])
        self.assertFalse(Project.objects.filter(project_no="H20240003").exists())

    def test_import_history_projects_rejects_non_finite_budget(self):
        workbook = DataCenterService.build_template("history_projects")
        sheet = workbook.active
        sheet.append(["H20240004", "非有限经费项目", "历史负责人", "计算机学院", "校级", "NaN", "2024"])
        stream = BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "history-non-finite-budget.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        task = DataCenterService.create_import_task(
            self.admin,
            "history_projects",
            uploaded,
        )
        finished = DataCenterService.run_import_task(task.id)

        self.assertEqual(finished.status, AsyncTaskRecord.TaskStatus.SUCCESS)
        self.assertEqual(finished.result["created"], 0)
        self.assertEqual(finished.result["errors"], ["Row 2: 经费必须为有效数字"])
        self.assertFalse(Project.objects.filter(project_no="H20240004").exists())

    def test_import_dictionary_items(self):
        workbook = DataCenterService.build_template("dictionaries")
        sheet = workbook.active
        sheet.append(["demo_type", "演示字典", "DEMO", "演示选项", "10"])
        stream = BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "dictionary.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        task = DataCenterService.create_import_task(self.admin, "dictionaries", uploaded)
        finished = DataCenterService.run_import_task(task.id)

        self.assertEqual(finished.status, AsyncTaskRecord.TaskStatus.SUCCESS)
        self.assertEqual(finished.result["created"], 1)
        item = DictionaryItem.objects.get(dict_type__code="demo_type", value="DEMO")
        self.assertEqual(item.label, "演示选项")
        self.assertEqual(item.sort_order, 10)
